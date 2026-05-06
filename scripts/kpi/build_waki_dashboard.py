r"""
TSA Waki KPI Dashboard — Final Version
3 indicators with data integrity analysis.
Uses fresh extraction from Google Sheets (_kpi_data_complete.json).
Output: ~/Downloads/RACCOONS_KPI_DASHBOARD.xlsx
"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import os
import json
import re
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(__file__)
DATA_PATH = os.path.join(SCRIPT_DIR, '..', '_kpi_data_complete.json')
OUTPUT_PATH = os.path.join(os.path.expanduser('~'), 'Downloads', 'RACCOONS_KPI_DASHBOARD_v2.xlsx')

# ─── Targets ───
T_ACCURACY = 0.90   # >90% within 1 week
T_DURATION = 28      # <4 weeks
T_RELIABILITY = 0.90 # >90% on-time

# ─── Styles ───
F_TITLE     = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
F_SUBTITLE  = Font(name='Segoe UI', size=10, color='A0A0A0')
F_H         = Font(name='Segoe UI', size=9, bold=True, color='FFFFFF')
F_B         = Font(name='Segoe UI', size=9, color='333333')
F_BB        = Font(name='Segoe UI', size=9, bold=True, color='333333')
F_KPI_BIG   = Font(name='Segoe UI', size=20, bold=True, color='1A1A2E')
F_KPI_LBL   = Font(name='Segoe UI', size=8, color='888888')
F_PASS      = Font(name='Segoe UI', size=10, bold=True, color='27AE60')
F_FAIL      = Font(name='Segoe UI', size=10, bold=True, color='E74C3C')
F_WARN      = Font(name='Segoe UI', size=10, bold=True, color='F39C12')
F_NA        = Font(name='Segoe UI', size=10, bold=True, color='95A5A6')
F_W         = Font(name='Segoe UI', size=9, bold=True, color='FFFFFF')
F_SECT      = Font(name='Segoe UI', size=11, bold=True, color='1A1A2E')
F_NOTE      = Font(name='Segoe UI', size=8, color='999999')
F_NOTE_I    = Font(name='Segoe UI', size=8, italic=True, color='999999')
F_CONF_H    = Font(name='Segoe UI', size=9, bold=True, color='27AE60')
F_CONF_M    = Font(name='Segoe UI', size=9, bold=True, color='F39C12')
F_CONF_L    = Font(name='Segoe UI', size=9, bold=True, color='E74C3C')

BG_DARK     = PatternFill('solid', fgColor='1A1A2E')
BG_HDR      = PatternFill('solid', fgColor='2C3E50')
BG_W        = PatternFill('solid', fgColor='FFFFFF')
BG_ALT      = PatternFill('solid', fgColor='F5F6F7')
BG_GREEN    = PatternFill('solid', fgColor='E8F5E9')
BG_RED      = PatternFill('solid', fgColor='FFEBEE')
BG_YELLOW   = PatternFill('solid', fgColor='FFF8E1')
BG_BLUE     = PatternFill('solid', fgColor='E3F2FD')
BG_GRAY     = PatternFill('solid', fgColor='F0F0F0')

AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
AR = Alignment(horizontal='right', vertical='center')
BD = Border(
    left=Side('thin', color='E0E0E0'), right=Side('thin', color='E0E0E0'),
    top=Side('thin', color='E0E0E0'), bottom=Side('thin', color='E0E0E0'),
)


# ─── Date parsing ───
def parse_date(raw):
    if not raw or raw.strip() in ('', '-', 'TBD', 'N/A'):
        return None
    raw = raw.strip().split('\n')[0].strip()  # first line only
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', raw)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})$', raw)
    if m:
        y = m.group(3)
        if len(y) == 2:
            y = '20' + y
        mo, day = int(m.group(1)), int(m.group(2))
        if mo > 12:
            mo, day = day, mo
        return f"{y}-{mo:02d}-{day:02d}"
    m = re.match(r'^(\d{1,2})-(\w+)$', raw)
    if m:
        day = int(m.group(1))
        mm = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
              'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12, 'jab': 1}
        mon = m.group(2).lower()
        if mon in mm:
            month = mm[mon]
            year = 2025 if month == 12 else 2026
            return f"{year}-{month:02d}-{day:02d}"
    return None


def first_eta(eta_raw):
    for line in eta_raw.split('\n'):
        d = parse_date(line.strip())
        if d:
            return d
    return None


def last_eta(eta_raw):
    dates = [parse_date(l.strip()) for l in eta_raw.split('\n')]
    dates = [d for d in dates if d]
    return max(dates) if dates else None


def days_between(d1, d2):
    return (datetime.strptime(d2, '%Y-%m-%d') - datetime.strptime(d1, '%Y-%m-%d')).days


def is_done(status):
    return 'done' in status.lower()


# ─── KPI Calculation ───
def analyze_person(tasks):
    """Full KPI analysis for one person. Returns dict with all metrics."""
    total = len(tasks)
    done = [t for t in tasks if is_done(t['status'])]

    # Done tasks with complete date data
    measured = []
    for t in done:
        eta = first_eta(t.get('eta', ''))
        delivery = parse_date(t.get('delivery_date', ''))
        date_add = parse_date(t.get('date_add', ''))
        if eta and delivery:
            delta = days_between(eta, delivery)
            dur = days_between(date_add, delivery) if date_add else None
            measured.append({
                'row': t['row'],
                'eta': eta,
                'delivery': delivery,
                'date_add': date_add,
                'delta': delta,
                'duration': dur,
                'trivial': eta == delivery,
                'customer': t.get('customer', '').strip(),
                'rescheduled': '\n' in t.get('eta', ''),
            })

    n = len(measured)
    trivial = sum(1 for m in measured if m['trivial'])
    non_trivial = [m for m in measured if not m['trivial']]

    # KPI 1: ETA Accuracy (within 1 week)
    within_1w = sum(1 for m in measured if abs(m['delta']) <= 7)
    kpi1 = within_1w / n if n else None

    # Non-trivial only (the REAL test of estimation)
    nt_within_1w = sum(1 for m in non_trivial if abs(m['delta']) <= 7)
    kpi1_real = nt_within_1w / len(non_trivial) if non_trivial else None

    # KPI 2: Duration (date_add -> delivery)
    durations = [m['duration'] for m in measured if m['duration'] is not None and m['duration'] >= 0]
    nonzero_dur = [d for d in durations if d > 0]
    kpi2_avg = sum(durations) / len(durations) if durations else None
    kpi2_med = sorted(durations)[len(durations) // 2] if durations else None
    kpi2_avg_nz = sum(nonzero_dur) / len(nonzero_dur) if nonzero_dur else None

    # Customer-level spans (first date_add -> last delivery per customer)
    cust_spans = {}
    for t in done:
        c = t.get('customer', '').strip()
        if not c:
            continue
        da = parse_date(t.get('date_add', ''))
        dl = parse_date(t.get('delivery_date', ''))
        if c not in cust_spans:
            cust_spans[c] = {'first_add': da, 'last_del': dl, 'count': 0, 'done': 0}
        if da and (not cust_spans[c]['first_add'] or da < cust_spans[c]['first_add']):
            cust_spans[c]['first_add'] = da
        if dl and (not cust_spans[c]['last_del'] or dl > cust_spans[c]['last_del']):
            cust_spans[c]['last_del'] = dl
        cust_spans[c]['count'] += 1
        cust_spans[c]['done'] += 1
    # Add non-done customer tasks
    for t in tasks:
        if is_done(t['status']):
            continue
        c = t.get('customer', '').strip()
        if c and c in cust_spans:
            cust_spans[c]['count'] += 1

    # KPI 3: Reliability (on-time rate)
    ontime = sum(1 for m in measured if m['delta'] <= 0)
    kpi3 = ontime / n if n else None

    # Non-trivial reliability
    nt_ontime = sum(1 for m in non_trivial if m['delta'] <= 0)
    kpi3_real = nt_ontime / len(non_trivial) if non_trivial else None

    # Data confidence score
    same_day_pct = sum(1 for d in durations if d == 0) / len(durations) if durations else 0
    trivial_pct = trivial / n if n else 0
    # High if: >70% measured, <30% same-day, <85% trivial
    # Medium if: >50% measured, <60% same-day
    # Low otherwise
    measured_pct = n / len(done) if done else 0
    if measured_pct >= 0.7 and same_day_pct < 0.30 and trivial_pct < 0.90:
        confidence = 'HIGH'
    elif measured_pct >= 0.5 and same_day_pct < 0.60:
        confidence = 'MEDIUM'
    else:
        confidence = 'LOW'

    # Rescheduled count
    rescheduled = sum(1 for m in measured if m['rescheduled'])

    return {
        'total': total,
        'done': len(done),
        'measured': n,
        'trivial': trivial,
        'non_trivial': len(non_trivial),
        # KPI 1
        'kpi1': kpi1,
        'kpi1_n': within_1w,
        'kpi1_real': kpi1_real,
        'kpi1_real_n': nt_within_1w,
        # KPI 2
        'kpi2_avg': kpi2_avg,
        'kpi2_med': kpi2_med,
        'kpi2_avg_nonzero': kpi2_avg_nz,
        'kpi2_durations': durations,
        'kpi2_same_day': sum(1 for d in durations if d == 0),
        'kpi2_under_4w': sum(1 for d in durations if d <= 28),
        'cust_spans': cust_spans,
        # KPI 3
        'kpi3': kpi3,
        'kpi3_n': ontime,
        'kpi3_late': n - ontime,
        'kpi3_real': kpi3_real,
        'kpi3_real_n': nt_ontime,
        'kpi3_real_late': len(non_trivial) - nt_ontime,
        # Data quality
        'confidence': confidence,
        'trivial_pct': trivial_pct,
        'same_day_pct': same_day_pct,
        'measured_pct': measured_pct,
        'rescheduled': rescheduled,
    }


# ─── Style helpers ───
def sc(cell, font=None, fill=None, alignment=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border


def bar(ws, row, text, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = BG_DARK
    cell = ws.cell(row=row, column=1, value=text)
    sc(cell, F_TITLE, BG_DARK, AL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 36


def sub(ws, row, text, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = BG_DARK
    cell = ws.cell(row=row, column=1, value=text)
    sc(cell, F_SUBTITLE, BG_DARK, AL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 18


def hdr_row(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        sc(cell, F_H, BG_HDR, AC, BD)
    ws.row_dimensions[row].height = 22


def data_row(ws, row, values, bold_col=0, fonts=None):
    rf = BG_W if row % 2 == 0 else BG_ALT
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        f = F_BB if c == bold_col else F_B
        if fonts and c in fonts:
            f = fonts[c]
        sc(cell, f, rf, AC, BD)
    if bold_col:
        ws.cell(row=row, column=bold_col).alignment = AL
    ws.row_dimensions[row].height = 20


def total_row(ws, row, values):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        sc(cell, F_W, BG_HDR, AC, BD)
    ws.cell(row=row, column=1).alignment = AL
    ws.row_dimensions[row].height = 24


def conf_font(conf):
    if conf == 'HIGH': return F_CONF_H
    if conf == 'MEDIUM': return F_CONF_M
    return F_CONF_L


def pf(val, target, higher=True):
    """Return (text, font, fill) for pass/fail."""
    if val is None:
        return 'N/A', F_NA, BG_GRAY
    ok = val >= target if higher else val <= target
    return ('PASS', F_PASS, BG_GREEN) if ok else ('FAIL', F_FAIL, BG_RED)


# ─── Tab 1: Executive Summary ───
def build_executive(wb, kpis):
    ws = wb.active
    ws.title = 'Executive'
    ws.sheet_properties.tabColor = '1A1A2E'
    NC = 10
    for i, w in enumerate([20, 14, 14, 14, 14, 14, 14, 14, 14, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    bar(ws, r, 'TSA Waki KPI Dashboard', NC); r += 1
    sub(ws, r, 'Period: Dec 2025 - Mar 2026 (Historical)  |  Source: TSA_Tasks_Consolidate  |  Generated: 2026-03-18', NC); r += 1
    r += 1  # blank

    # ── KPI 1 ──
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).fill = BG_BLUE
    cell = ws.cell(row=r, column=1, value='KPI 1: ETA Accuracy')
    sc(cell, F_SECT, BG_BLUE, AL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    cell2 = ws.cell(row=r, column=8, value='Target: >90% within 1 week of original ETA')
    sc(cell2, Font(name='Segoe UI', size=8, bold=True, color='2C3E50'), BG_BLUE, AR)
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=NC)
    ws.row_dimensions[r].height = 26; r += 1

    hdr_row(ws, r, ['Person', 'Done', 'Measurable', 'Within 1w', 'Accuracy', 'Status',
                     'Trivial (ETA=Del)', 'Real Tests', 'Real Accuracy', 'Confidence']); r += 1

    names = list(kpis.keys())
    for name in names:
        k = kpis[name]
        st, sf, sbg = pf(k['kpi1'], T_ACCURACY)
        # Real accuracy status
        rst, rsf, rsbg = pf(k['kpi1_real'], T_ACCURACY) if k['non_trivial'] >= 3 else ('N/S', F_NA, BG_GRAY)
        vals = [
            name, k['done'], k['measured'],
            k['kpi1_n'], f"{k['kpi1']*100:.0f}%" if k['kpi1'] is not None else 'N/A',
            st,
            f"{k['trivial']} ({k['trivial_pct']*100:.0f}%)",
            k['non_trivial'],
            f"{k['kpi1_real']*100:.0f}%" if k['kpi1_real'] is not None else 'N/S',
            k['confidence'],
        ]
        fonts_map = {6: sf, 9: rsf, 10: conf_font(k['confidence'])}
        fills = {6: sbg, 9: rsbg}
        rf = BG_W if r % 2 == 0 else BG_ALT
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            fnt = fonts_map.get(c, F_BB if c == 1 else F_B)
            fll = fills.get(c, rf)
            sc(cell, fnt, fll, AC, BD)
        ws.cell(row=r, column=1).alignment = AL
        ws.row_dimensions[r].height = 20; r += 1

    # Team total
    t_measured = sum(k['measured'] for k in kpis.values())
    t_w1w = sum(k['kpi1_n'] for k in kpis.values())
    t_nt = sum(k['non_trivial'] for k in kpis.values())
    t_nt_w1w = sum(k['kpi1_real_n'] for k in kpis.values())
    t_acc = t_w1w / t_measured if t_measured else None
    t_racc = t_nt_w1w / t_nt if t_nt else None
    st, _, _ = pf(t_acc, T_ACCURACY)
    rst, _, _ = pf(t_racc, T_ACCURACY) if t_nt >= 5 else ('N/S', None, None)
    total_row(ws, r, [
        'TEAM', sum(k['done'] for k in kpis.values()), t_measured, t_w1w,
        f"{t_acc*100:.0f}%" if t_acc else 'N/A', st,
        sum(k['trivial'] for k in kpis.values()), t_nt,
        f"{t_racc*100:.0f}%" if t_racc else 'N/S', ''
    ]); r += 2

    # ── KPI 2 ──
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).fill = BG_YELLOW
    cell = ws.cell(row=r, column=1, value='KPI 2: Implementation Velocity')
    sc(cell, F_SECT, BG_YELLOW, AL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    cell2 = ws.cell(row=r, column=8, value='Target: <4 weeks (28 days) from start to delivery')
    sc(cell2, Font(name='Segoe UI', size=8, bold=True, color='2C3E50'), BG_YELLOW, AR)
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=NC)
    ws.row_dimensions[r].height = 26; r += 1

    hdr_row(ws, r, ['Person', 'Tasks w/ Dur', 'Avg (all)', 'Median', 'Max',
                     'Under 4w', 'Status', 'Same-day', 'Avg (excl 0d)', 'Confidence']); r += 1

    for name in names:
        k = kpis[name]
        d = k['kpi2_durations']
        st, sf, sbg = pf(k['kpi2_avg'], T_DURATION, higher=False) if k['kpi2_avg'] is not None else ('N/A', F_NA, BG_GRAY)
        u4w = k['kpi2_under_4w']
        u4w_pct = f"{u4w/len(d)*100:.0f}%" if d else 'N/A'
        vals = [
            name, len(d),
            f"{k['kpi2_avg']:.1f}d" if k['kpi2_avg'] is not None else 'N/A',
            f"{k['kpi2_med']}d" if k['kpi2_med'] is not None else 'N/A',
            f"{max(d)}d" if d else 'N/A',
            f"{u4w} ({u4w_pct})",
            st,
            f"{k['kpi2_same_day']} ({k['same_day_pct']*100:.0f}%)",
            f"{k['kpi2_avg_nonzero']:.1f}d" if k['kpi2_avg_nonzero'] is not None else 'N/A',
            k['confidence'],
        ]
        fonts_map = {7: sf, 10: conf_font(k['confidence'])}
        fills = {7: sbg}
        rf = BG_W if r % 2 == 0 else BG_ALT
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            fnt = fonts_map.get(c, F_BB if c == 1 else F_B)
            fll = fills.get(c, rf)
            sc(cell, fnt, fll, AC, BD)
        ws.cell(row=r, column=1).alignment = AL
        ws.row_dimensions[r].height = 20; r += 1

    all_dur = []
    for k in kpis.values():
        all_dur.extend(k['kpi2_durations'])
    t_avg = sum(all_dur) / len(all_dur) if all_dur else None
    t_med = sorted(all_dur)[len(all_dur) // 2] if all_dur else None
    nz = [d for d in all_dur if d > 0]
    st, _, _ = pf(t_avg, T_DURATION, higher=False) if t_avg is not None else ('N/A', None, None)
    total_row(ws, r, [
        'TEAM', len(all_dur),
        f"{t_avg:.1f}d" if t_avg else 'N/A',
        f"{t_med}d" if t_med is not None else 'N/A',
        f"{max(all_dur)}d" if all_dur else 'N/A',
        f"{sum(1 for d in all_dur if d <= 28)} ({sum(1 for d in all_dur if d <= 28)/len(all_dur)*100:.0f}%)" if all_dur else '',
        st,
        sum(k['kpi2_same_day'] for k in kpis.values()),
        f"{sum(nz)/len(nz):.1f}d" if nz else 'N/A', ''
    ]); r += 2

    # ── KPI 3 ──
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).fill = BG_GREEN
    cell = ws.cell(row=r, column=1, value='KPI 3: Implementation Reliability')
    sc(cell, F_SECT, BG_GREEN, AL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    cell2 = ws.cell(row=r, column=8, value='Target: >90% delivered on or before ETA')
    sc(cell2, Font(name='Segoe UI', size=8, bold=True, color='2C3E50'), BG_GREEN, AR)
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=NC)
    ws.row_dimensions[r].height = 26; r += 1

    hdr_row(ws, r, ['Person', 'Measurable', 'On Time', 'Late', 'Rate', 'Status',
                     'Trivial (ETA=Del)', 'Real Tests', 'Real Rate', 'Confidence']); r += 1

    for name in names:
        k = kpis[name]
        st, sf, sbg = pf(k['kpi3'], T_RELIABILITY)
        rst, rsf, rsbg = pf(k['kpi3_real'], T_RELIABILITY) if k['non_trivial'] >= 3 else ('N/S', F_NA, BG_GRAY)
        vals = [
            name, k['measured'], k['kpi3_n'], k['kpi3_late'],
            f"{k['kpi3']*100:.0f}%" if k['kpi3'] is not None else 'N/A', st,
            f"{k['trivial']} ({k['trivial_pct']*100:.0f}%)",
            k['non_trivial'],
            f"{k['kpi3_real']*100:.0f}%" if k['kpi3_real'] is not None else 'N/S',
            k['confidence'],
        ]
        fonts_map = {6: sf, 9: rsf, 10: conf_font(k['confidence'])}
        fills = {6: sbg, 9: rsbg}
        rf = BG_W if r % 2 == 0 else BG_ALT
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            fnt = fonts_map.get(c, F_BB if c == 1 else F_B)
            fll = fills.get(c, rf)
            sc(cell, fnt, fll, AC, BD)
        ws.cell(row=r, column=1).alignment = AL
        ws.row_dimensions[r].height = 20; r += 1

    t_ontime = sum(k['kpi3_n'] for k in kpis.values())
    t_late = sum(k['kpi3_late'] for k in kpis.values())
    t_rel = t_ontime / t_measured if t_measured else None
    t_nt_ot = sum(k['kpi3_real_n'] for k in kpis.values())
    t_nt_late = sum(k['kpi3_real_late'] for k in kpis.values())
    t_rrel = t_nt_ot / t_nt if t_nt else None
    st, _, _ = pf(t_rel, T_RELIABILITY)
    total_row(ws, r, [
        'TEAM', t_measured, t_ontime, t_late,
        f"{t_rel*100:.0f}%" if t_rel else 'N/A', st,
        sum(k['trivial'] for k in kpis.values()), t_nt,
        f"{t_rrel*100:.0f}%" if t_rrel else 'N/S', ''
    ]); r += 2

    # ── Interpretation ──
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).fill = BG_GRAY
    cell = ws.cell(row=r, column=1, value='How to Read This Dashboard')
    sc(cell, F_SECT, BG_GRAY, AL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    ws.row_dimensions[r].height = 22; r += 1

    notes = [
        '"Trivial" = tasks where ETA and Delivery Date are identical. These are on-time by definition but don\'t test estimation quality.',
        '"Real Tests" = tasks where ETA differs from Delivery. These are the genuine tests of estimation accuracy.',
        '"Real Accuracy/Rate" only shows when there are 3+ real tests. "N/S" = Not Significant (too few data points).',
        '"Same-day" in KPI 2 = tasks completed same day as added. High % may indicate retroactive logging.',
        '"Avg (excl 0d)" removes same-day tasks for a more realistic velocity picture.',
        'Confidence: HIGH = robust data, LOW = many same-day or trivial entries — interpret with caution.',
        'Data source: Google Sheets (Dec 2025 - Mar 2026). Moving to Linear will provide real-time tracking.',
        'Diego is excluded (already tracked directly in Linear, not migrated from Sheets).',
    ]
    for note in notes:
        cell = ws.cell(row=r, column=1, value=f"  {note}")
        sc(cell, F_NOTE, BG_W, AL)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        ws.row_dimensions[r].height = 14; r += 1

    ws.freeze_panes = 'A4'


# ─── Tab 2: Data Integrity ───
def build_integrity(wb, kpis):
    ws = wb.create_sheet('Data Integrity')
    ws.sheet_properties.tabColor = 'E74C3C'
    NC = 8
    for i, w in enumerate([20, 14, 14, 14, 14, 14, 14, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    bar(ws, r, 'Data Integrity Analysis', NC); r += 1
    sub(ws, r, 'Understanding data quality before trusting the numbers', NC); r += 2

    # Data coverage
    cell = ws.cell(row=r, column=1, value='Data Coverage (Done tasks with complete date fields)')
    sc(cell, F_SECT, BG_BLUE, AL)
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).fill = BG_BLUE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    ws.row_dimensions[r].height = 26; r += 1

    hdr_row(ws, r, ['Person', 'Total', 'Done', 'Measurable', 'Coverage', 'Missing ETA', 'Missing Delivery', 'Confidence']); r += 1

    for name, k in kpis.items():
        missing_del = k['done'] - k['measured']
        missing_eta = k['done'] - k['measured']  # simplified
        vals = [
            name, k['total'], k['done'], k['measured'],
            f"{k['measured_pct']*100:.0f}%",
            '', f"{missing_del}",
            k['confidence'],
        ]
        fonts_map = {8: conf_font(k['confidence'])}
        data_row(ws, r, vals, bold_col=1, fonts=fonts_map); r += 1
    r += 1

    # Logging patterns
    cell = ws.cell(row=r, column=1, value='Logging Pattern Analysis')
    sc(cell, F_SECT, BG_YELLOW, AL)
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).fill = BG_YELLOW
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    ws.row_dimensions[r].height = 26; r += 1

    hdr_row(ws, r, ['Person', 'Measured', 'ETA=Delivery', '% Trivial', 'Same-day Dur', '% Same-day', 'Rescheduled', 'Pattern']); r += 1

    for name, k in kpis.items():
        pattern = 'Real-time tracking' if k['same_day_pct'] < 0.30 and k['trivial_pct'] < 0.90 else \
                  'Mixed logging' if k['same_day_pct'] < 0.60 else 'Likely retroactive'
        vals = [
            name, k['measured'], k['trivial'],
            f"{k['trivial_pct']*100:.0f}%",
            k['kpi2_same_day'],
            f"{k['same_day_pct']*100:.0f}%",
            k['rescheduled'],
            pattern,
        ]
        pf_map = {
            8: F_CONF_H if 'Real-time' in pattern else (F_CONF_M if 'Mixed' in pattern else F_CONF_L)
        }
        data_row(ws, r, vals, bold_col=1, fonts=pf_map); r += 1
    r += 1

    # Assessment
    cell = ws.cell(row=r, column=1, value='Assessment')
    sc(cell, F_SECT, BG_GRAY, AL)
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).fill = BG_GRAY
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    ws.row_dimensions[r].height = 24; r += 1

    assessments = [
        'The spreadsheet was used as a RECORD of what happened, not as a real-time estimation tracker.',
        '85-96% of delivered tasks have identical ETA and Delivery dates — this means most entries are logged AFTER the fact.',
        'Carlos has the most informative data: 15% trivial gap, 15 rescheduled ETAs, realistic durations.',
        'Thiago has 80% same-day completion and 95% trivial — his metrics look perfect but reflect logging patterns.',
        'Gabi has 60% same-day tasks — mixed pattern, metrics directionally correct but not precise.',
        'Alexandra has 12% same-day and consistent dates — second most reliable data after Carlos.',
        '',
        'RECOMMENDATION: Linear will fix this. Use Linear due dates for ETA, completedAt for delivery.',
        'These historical KPIs are directionally positive but insufficient for SLA-level commitments.',
        'For customer-facing promises, use Carlos-level rigor as the benchmark going forward.',
    ]
    for a in assessments:
        cell = ws.cell(row=r, column=1, value=f"  {a}" if a else '')
        sc(cell, F_NOTE_I if a.startswith('REC') else F_NOTE, BG_W, AL)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        ws.row_dimensions[r].height = 14 if a else 8; r += 1


# ─── Tab 3: Customer View ───
def build_customer(wb, kpis, raw_data):
    ws = wb.create_sheet('By Customer')
    ws.sheet_properties.tabColor = '27AE60'
    NC = 8
    for i, w in enumerate([18, 16, 10, 10, 12, 16, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    bar(ws, r, 'Customer Implementation Spans', NC); r += 1
    sub(ws, r, 'From first task to last delivery per customer — proxy for KPI 2 (4-week target)', NC); r += 2

    hdr_row(ws, r, ['Customer', 'Owner', 'Tasks', 'Done', 'First Task', 'Last Delivery', 'Span (days)', 'Under 4w?']); r += 1

    person_map = {'alexandra': 'Alexandra', 'gabi': 'Gabrielle', 'carlos': 'Carlos', 'thiago': 'Thiago'}
    all_spans = []

    for person_key, person_name in person_map.items():
        k = kpis[person_name]
        for cust, info in sorted(k['cust_spans'].items(), key=lambda x: x[1]['done'], reverse=True):
            if not info['first_add'] or not info['last_del']:
                continue
            try:
                span = days_between(info['first_add'], info['last_del'])
            except Exception:
                continue
            if span < 0:
                continue
            all_spans.append(span)
            under = 'Yes' if span <= 28 else 'No'
            st_font = F_PASS if span <= 28 else (F_WARN if span <= 42 else F_FAIL)
            vals = [
                cust, person_name, info['count'], info['done'],
                info['first_add'], info['last_del'],
                f"{span}d", under,
            ]
            fonts_map = {8: st_font}
            data_row(ws, r, vals, bold_col=1, fonts=fonts_map); r += 1

    r += 1
    # Summary
    if all_spans:
        avg_span = sum(all_spans) / len(all_spans)
        med_span = sorted(all_spans)[len(all_spans) // 2]
        under_4w = sum(1 for s in all_spans if s <= 28)
        cell = ws.cell(row=r, column=1, value=f"Summary: {len(all_spans)} customer engagements | "
                       f"Avg span: {avg_span:.0f}d | Median: {med_span}d | "
                       f"Under 4 weeks: {under_4w}/{len(all_spans)} ({under_4w/len(all_spans)*100:.0f}%)")
        sc(cell, F_BB, BG_GRAY, AL)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        ws.row_dimensions[r].height = 22; r += 1

    r += 1
    cell = ws.cell(row=r, column=1,
                   value="  Note: Spans include ALL task types (not just onboarding). "
                         "Ongoing customers show full engagement duration, not just initial implementation.")
    sc(cell, F_NOTE, BG_W, AL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)


# ─── Main ───
def main():
    print('=' * 60)
    print('BUILDING WAKI KPI DASHBOARD (Final Version)')
    print('=' * 60)

    with open(DATA_PATH, 'r', encoding='utf-8') as f:
        raw = json.load(f)

    print('\nAnalyzing...')
    kpis = {}
    for key, display in [('alexandra', 'Alexandra'), ('gabi', 'Gabrielle'),
                          ('carlos', 'Carlos'), ('thiago', 'Thiago')]:
        k = analyze_person(raw[key])
        kpis[display] = k
        acc = f"{k['kpi1']*100:.0f}%" if k['kpi1'] else 'N/A'
        rel = f"{k['kpi3']*100:.0f}%" if k['kpi3'] else 'N/A'
        dur = f"{k['kpi2_avg']:.1f}d" if k['kpi2_avg'] is not None else 'N/A'
        print(f"  {display}: {k['done']} done, {k['measured']} measured | "
              f"Acc={acc} Rel={rel} Dur={dur} | "
              f"Trivial={k['trivial_pct']*100:.0f}% SameDay={k['same_day_pct']*100:.0f}% | "
              f"Conf={k['confidence']}")

    print('\nBuilding XLSX...')
    wb = Workbook()
    build_executive(wb, kpis)
    print('  Tab 1: Executive (3 KPIs + interpretation)')
    build_integrity(wb, kpis)
    print('  Tab 2: Data Integrity (quality analysis)')
    build_customer(wb, kpis, raw)
    print('  Tab 3: By Customer (implementation spans)')

    wb.save(OUTPUT_PATH)
    print(f'\nSaved: {OUTPUT_PATH}')
    print('=' * 60)


if __name__ == '__main__':
    main()
