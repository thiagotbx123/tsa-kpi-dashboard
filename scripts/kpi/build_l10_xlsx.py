"""
L10 Scorecard XLSX for Waki — 3 dashboard indicators:
  1. Customer Onboarding (per-customer, from implementation_timeline.json)
  2. ETA Accuracy (per-TSA per-period, "TSA KPI")
  3. Implementation Reliability (per-TSA per-period, rework)

Reads _dashboard_data.json + implementation_timeline.json.
Writes ~/Downloads/L10_SCORECARD_WAKI.xlsx.
"""
import json
import os
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
DATA = SCRIPT_DIR.parent / "_dashboard_data.json"
TIMELINE_PATH = SCRIPT_DIR / "implementation_timeline.json"
OUT = Path.home() / "Downloads" / "L10_SCORECARD_WAKI.xlsx"

KPI_MEMBERS = ["THIAGO", "CARLOS", "ALEXANDRA", "DIEGO", "GABI"]

# KPI thresholds (red / yellow / green)
KPI1_GREEN, KPI1_YELLOW = 0.90, 0.75   # ETA Accuracy
KPI3_GREEN, KPI3_YELLOW = 0.90, 0.75   # Reliability
# Customer Onboarding: months from kickoff → go-live
ONB_GREEN, ONB_YELLOW = 3, 6           # green ≤ 3mo, yellow ≤ 6mo
# Current TSA team era cutoff (mirrors dashboard line 3157)
TSA_ERA_YEARMONTH = 202508             # Sep/25 (0-indexed month 8)

# Dashboard parity: replicate getKPIFiltered() from build_html_dashboard.py
#   - drop spreadsheet source
#   - external category only
#   - rolling 4 months back from today (isCoreWeek)
CORE_WEEK_ROLLING_MONTHS = 4


def _core_week_cutoff():
    t = date.today()
    m = t.month - CORE_WEEK_ROLLING_MONTHS
    y = t.year
    while m <= 0:
        m += 12
        y -= 1
    return date(y, m, 1)


def _parse_week(w):
    try:
        yy, rest = w.split("-", 1)
        mm, wpart = rest.split(" ", 1)
        return int(yy), int(mm), int(wpart.replace("W.", ""))
    except Exception:
        return None


def is_core_week(w, cutoff, today):
    p = _parse_week(w)
    if not p:
        return False
    yy, mm, wn = p
    wy = 2000 + yy if yy < 50 else 1900 + yy
    try:
        wdate = date(wy, mm, min((wn - 1) * 7 + 1, 28))
    except ValueError:
        return False
    if wdate > today:
        return False
    return wdate >= cutoff


def filter_kpi(records):
    """Replicate dashboard getKPIFiltered(): Linear + External + core weeks + KPI members."""
    today = date.today()
    cutoff = _core_week_cutoff()
    out = []
    for r in records:
        if r.get("source") == "spreadsheet":
            continue
        if r.get("category") != "External":
            continue
        if r.get("tsa") not in KPI_MEMBERS:
            continue
        w = r.get("week", "")
        if not w or not is_core_week(w, cutoff, today):
            continue
        out.append(r)
    return out

# Colors
C_HEADER = PatternFill("solid", fgColor="1F3864")
C_SUB = PatternFill("solid", fgColor="2E5597")
C_GREEN = PatternFill("solid", fgColor="C6EFCE")
C_YELLOW = PatternFill("solid", fgColor="FFEB9C")
C_RED = PatternFill("solid", fgColor="FFC7CE")
C_GRAY = PatternFill("solid", fgColor="E7E6E6")
C_CARD = PatternFill("solid", fgColor="F2F2F2")

F_TITLE = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
F_H = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_SUB = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_BODY = Font(name="Calibri", size=10)
F_BODY_B = Font(name="Calibri", size=10, bold=True)
F_SMALL = Font(name="Calibri", size=9, italic=True, color="595959")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def load():
    with open(DATA, "r", encoding="utf-8") as f:
        return json.load(f)


def load_timeline():
    with open(TIMELINE_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


MON_MAP = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
           "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}


def parse_mon_yy(s):
    """'Mar/22' -> date(2022,3,1). Returns None for invalid."""
    if not s:
        return None
    try:
        mon_str, yy = s.split("/")
        return date(2000 + int(yy), MON_MAP[mon_str], 1)
    except Exception:
        return None


def months_diff(a, b):
    if not (a and b):
        return None
    return (b.year - a.year) * 12 + (b.month - a.month)


def yearmonth_num(d):
    """date -> int YYYYMM for comparison."""
    return d.year * 100 + (d.month - 1)  # 0-indexed month like dashboard


def yy_to_yyyy(yy):
    yy = int(yy)
    return 2000 + yy if yy >= 20 else 2000 + yy


def period_from_week(week_str):
    """'26-03 W.2' -> '2026-03'"""
    if not week_str or "-" not in week_str:
        return ""
    try:
        yy, rest = week_str.split("-", 1)
        mm = rest.split(" ", 1)[0]
        return f"{yy_to_yyyy(yy):04d}-{int(mm):02d}"
    except Exception:
        return ""


def parse_date(s):
    if not s or not isinstance(s, str):
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ"):
        try:
            return datetime.strptime(s[:len(fmt.replace('%f','000000'))], fmt).date()
        except Exception:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except Exception:
        return None


# ---------------- KPI computations ----------------

def compute_kpi1(records):
    """ETA Accuracy: On Time / (On Time + Late) per (TSA, period).
    Records are expected already filtered by filter_kpi()."""
    buckets = defaultdict(lambda: {"on_time": 0, "late": 0})
    for r in records:
        perf = r.get("perf", "")
        if perf not in ("On Time", "Late"):
            continue
        period = period_from_week(r.get("week", ""))
        if not period:
            continue
        tsa = r.get("tsa", "")
        if perf == "On Time":
            buckets[(tsa, period)]["on_time"] += 1
        else:
            buckets[(tsa, period)]["late"] += 1
    return buckets


def compute_kpi2(records):
    """Velocity: avg(delivery - startedAt|dateAdd) days per (TSA, period).

    Mirrors dashboard calcVelocity(): uses r.delivery + (r.startedAt || r.dateAdd),
    status=Done only, skips negative durations. No upper cap — matches dashboard.
    """
    buckets = defaultdict(list)
    for r in records:
        if r.get("status", "") != "Done":
            continue
        start_raw = r.get("startedAt", "") or r.get("dateAdd", "")
        delivery_raw = r.get("delivery", "")
        sd = parse_date(start_raw)
        dd = parse_date(delivery_raw)
        if not (sd and dd):
            continue
        days = (dd - sd).days
        if days < 0:
            continue
        period = f"{dd.year:04d}-{dd.month:02d}"
        buckets[(r.get("tsa", ""), period)].append(days)
    return buckets


def compute_kpi3(records):
    """Reliability: Done without Rework / Total Done per (TSA, period).
    Records are expected already filtered by filter_kpi()."""
    buckets = defaultdict(lambda: {"done": 0, "rework": 0})
    for r in records:
        if r.get("status", "") != "Done":
            continue
        period = period_from_week(r.get("week", ""))
        if not period:
            continue
        key = (r.get("tsa", ""), period)
        buckets[key]["done"] += 1
        if r.get("rework") == "yes":
            buckets[key]["rework"] += 1
    return buckets


def periods_sorted(*bucket_dicts, n_last=12):
    today = date.today()
    cutoff = f"{today.year:04d}-{today.month:02d}"
    s = set()
    for b in bucket_dicts:
        for (_, p) in b.keys():
            if p and p <= cutoff:
                s.add(p)
    out = sorted(s)
    return out[-n_last:] if len(out) > n_last else out


# ---------------- Writers ----------------

def write_title(ws, title, cols_merge):
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(cols_merge)}1")
    c = ws["A1"]
    c.value = title
    c.fill = C_HEADER
    c.font = F_TITLE
    c.alignment = CENTER


def write_cover(wb):
    ws = wb.create_sheet("L10 Reading Guide", 0)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 95

    write_title(ws, "L10 Scorecard — TSA Raccoons — Quick Reading Guide", 2)

    rows = [
        ("", ""),
        ("What is L10", "Level 10 Meeting (EOS): weekly 90-min meeting. "
         "The Scorecard Review block (5 min) goes through each KPI: each owner says 'on track' or 'off track', "
         "no discussion — issues go to the Issues List. Every metric has an OWNER + WEEKLY TARGET + Green/Yellow/Red thresholds."),
        ("How we read this", "We track the 3 indicators the TSA dashboard already exposes: Customer Onboarding, "
         "ETA Accuracy (TSA KPI) and Implementation Reliability (rework). This XLSX MIRRORS the live dashboard "
         "(KPI_DASHBOARD.html): ETA/Reliability use Linear + External(Customer) + rolling 4-month core window; "
         "Customer Onboarding uses the implementation_timeline dataset (18 projects with kickoff/go-live)."),
        ("", ""),
        ("Indicator 1 — Customer Onboarding", "Time from project kickoff to go-live (months) per customer. "
         "Captures implementation speed at project level, not ticket level. Uses implementation_timeline.json "
         "(the same data powering the 'Customer Onboarding' dashboard tab). TSA Team Avg = projects kicked off "
         "Sep/25 onwards (current team era)."),
        ("Indicator 1 Target", "Green ≤ 3 months   |   Yellow 4–6 months   |   Red > 6 months"),
        ("", ""),
        ("Indicator 2 — ETA Accuracy (TSA KPI)", "On Time / (On Time + Late). Measures whether the TSA delivers within the due date committed. "
         "Excludes 'Not Started', 'No ETA', 'On Hold', 'Blocked'. B.B.C. (Blocked By Customer) does not count as Late. "
         "Admin-closed = On Time (D.LIE15). Canceled = N/A. "
         "D.LIE26: 'No ETA' only counts for External(Customer) demand — Internal/non-customer tickets without ETA go to N/A. "
         "D.LIE27: On Time = delivered within dueDate + 1 business day (weekend-aware tolerance buffer)."),
        ("Indicator 2 Target", "Green ≥ 90%   |   Yellow 75–89%   |   Red < 75%"),
        ("", ""),
        ("Indicator 3 — Implementation Reliability", "Done WITHOUT Rework / Total Done. Rework = ticket with the "
         "'rework:implementation' label in Linear (or Done → In Progress transition, D.LIE20). Captures delivery quality."),
        ("Indicator 3 Target", "Green ≥ 90%   |   Yellow 75–89%   |   Red < 75%"),
        ("", ""),
        ("KPI Onboarding (team)", "A TSA joins KPI when operating in Raccoons Linear flow. "
         "Thais/Yasmim removed from KPI on 2026-04-09 (cross-team Opossum). Samples <10/month are flagged low sample."),
        ("", ""),
        ("Data integrity rules (key D.LIE)", "D.LIE10 B.B.C. excluded from Overdue • D.LIE14 deliveryDate activity-based • "
         "D.LIE15 Admin-closed = On Time • D.LIE19 Parent tickets excluded • D.LIE20 Rework = Done→In Progress • "
         "D.LIE22 Linear supersedes Sheets • D.LIE23 Original assignee owns the ticket • "
         "D.LIE26 No ETA only for External(Customer) • D.LIE27 1 business-day tolerance on dueDate."),
        ("", ""),
        ("Data sources", "Linear GraphQL API (KPI2/KPI3) + implementation_timeline.json (KPI1). "
         "Live HTML dashboard: ~/Downloads/KPI_DASHBOARD.html (pipeline at Tools/TSA_CORTEX/scripts/kpi/)."),
        ("Note — Execution Time / Velocity", "The dashboard has an 'Execution Time' tab (internally 'velocity'). "
         "NOT included here — it is not one of the 3 core L10 indicators. Check the live dashboard if needed."),
        ("Build", f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')} | sources: _dashboard_data.json + implementation_timeline.json"),
    ]

    for i, (k, v) in enumerate(rows, start=2):
        ws.row_dimensions[i].height = 38 if v and len(v) > 80 else 20
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v)
        a.font = F_BODY_B
        a.alignment = LEFT
        b.font = F_BODY
        b.alignment = LEFT
        if k and k.startswith("KPI"):
            a.fill = C_SUB
            a.font = F_SUB
            b.fill = C_CARD
        elif k.startswith("Meta"):
            a.fill = C_GRAY
            a.font = F_BODY_B


def write_resumo(wb, k1, k3, periods, timeline):
    ws = wb.create_sheet("Summary", 1)
    cols = ["TSA / Scope", "ETA %", "ETA Sample", "ETA Status",
            "Reliability %", "Rel Sample", "Rel Status"]
    span = f"{periods[0]} → {periods[-1]}" if periods else "no data"
    write_title(ws, f"Consolidated summary by TSA — period {span}", len(cols))

    # header
    for j, h in enumerate(cols, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.fill = C_SUB
        c.font = F_H
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[3].height = 30

    pset = set(periods)
    # Portfolio block (Customer Onboarding) — row 4
    completed = []
    for r in timeline:
        kd = parse_mon_yy(r.get("kickoff"))
        gd = parse_mon_yy(r.get("goLive"))
        m = months_diff(kd, gd) if kd and gd else None
        if m is not None:
            completed.append((r.get("customer"), m, r.get("status"), yearmonth_num(kd) if kd else 0))
    avg_months = sum(x[1] for x in completed) / len(completed) if completed else None
    tsa_era = [x for x in completed if x[3] >= TSA_ERA_YEARMONTH]
    tsa_avg = sum(x[1] for x in tsa_era) / len(tsa_era) if tsa_era else None
    live_maint = sum(1 for r in timeline if r.get("status") in ("live", "maintenance"))

    ws.cell(row=4, column=1, value="Customer Onboarding (portfolio)").font = F_BODY_B
    ws.cell(row=4, column=1).fill = C_CARD
    ws.cell(row=4, column=2, value=f"avg {avg_months:.1f} mo" if avg_months else "—").font = F_BODY
    ws.cell(row=4, column=3, value=f"{len(completed)} projects").font = F_BODY
    ws.cell(row=4, column=4, value=farol_low(avg_months, ONB_GREEN, ONB_YELLOW) if avg_months else "—")
    paint_farol(ws.cell(row=4, column=4), farol_low(avg_months, ONB_GREEN, ONB_YELLOW) if avg_months else "—")
    ws.cell(row=4, column=5, value=f"TSA era avg: {tsa_avg:.1f} mo" if tsa_avg else "—").font = F_BODY
    ws.cell(row=4, column=6, value=f"{len(tsa_era)} projects").font = F_BODY
    ws.cell(row=4, column=7, value=farol_low(tsa_avg, ONB_GREEN, ONB_YELLOW) if tsa_avg else "—")
    paint_farol(ws.cell(row=4, column=7), farol_low(tsa_avg, ONB_GREEN, ONB_YELLOW) if tsa_avg else "—")
    for col in range(1, 8):
        ws.cell(row=4, column=col).border = BORDER
        ws.cell(row=4, column=col).alignment = CENTER

    # Spacer row
    ws.cell(row=5, column=1, value="Per-TSA — ETA Accuracy & Reliability (indicators 2 & 3)").font = F_BODY_B
    ws.cell(row=5, column=1).alignment = LEFT

    team_ot = team_lt = team_done = team_rew = 0

    for i, tsa in enumerate(KPI_MEMBERS, start=6):
        # ETA
        on_time = sum(v["on_time"] for (t, p), v in k1.items() if t == tsa and p in pset)
        late = sum(v["late"] for (t, p), v in k1.items() if t == tsa and p in pset)
        n1 = on_time + late
        v1 = (on_time / n1) if n1 else None
        team_ot += on_time
        team_lt += late

        # Reliability
        done = sum(v["done"] for (t, p), v in k3.items() if t == tsa and p in pset)
        rew = sum(v["rework"] for (t, p), v in k3.items() if t == tsa and p in pset)
        n3 = done
        v3 = ((done - rew) / done) if done else None
        team_done += done
        team_rew += rew

        row = [tsa,
               v1, n1, farol_pct(v1, KPI1_GREEN, KPI1_YELLOW),
               v3, n3, farol_pct(v3, KPI3_GREEN, KPI3_YELLOW)]

        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = F_BODY_B if j == 1 else F_BODY
            c.alignment = CENTER
            c.border = BORDER
            if j in (2, 5) and isinstance(val, float):
                c.number_format = "0.0%"
            if j in (4, 7):
                paint_farol(c, val)

    # TEAM row — weighted mean (same as dashboard)
    team_n1 = team_ot + team_lt
    team_eta = (team_ot / team_n1) if team_n1 else None
    team_rel = ((team_done - team_rew) / team_done) if team_done else None
    team_row_i = 6 + len(KPI_MEMBERS)

    team_row = ["TEAM (weighted — dashboard)",
                team_eta, team_n1, farol_pct(team_eta, KPI1_GREEN, KPI1_YELLOW),
                team_rel, team_done, farol_pct(team_rel, KPI3_GREEN, KPI3_YELLOW)]

    for j, val in enumerate(team_row, start=1):
        c = ws.cell(row=team_row_i, column=j, value=val)
        c.font = F_BODY_B
        c.alignment = LEFT if j == 1 else CENTER
        c.border = BORDER
        c.fill = C_CARD
        if j in (2, 5) and isinstance(val, float):
            c.number_format = "0.0%"
        if j in (4, 7):
            paint_farol(c, val)

    # Explanatory note
    note_row = team_row_i + 2
    ws.cell(row=note_row, column=1,
            value=("NOTE: TEAM % = weighted mean = sum(on_time across TSAs) / sum(on_time+late). "
                   "This matches the dashboard. Simple average of the 5 TSA rows above gives a "
                   "different number (89.8%) because Diego's 14-ticket sample gets the same weight "
                   "as Carlos's 145-ticket sample — statistically misleading."))
    ws.cell(row=note_row, column=1).alignment = LEFT
    ws.cell(row=note_row, column=1).font = F_SMALL
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
    ws.row_dimensions[note_row].height = 42

    # widths
    ws.column_dimensions["A"].width = 32
    for L in "BCDEFG":
        ws.column_dimensions[L].width = 18


def farol_pct(v, green, yellow):
    if v is None:
        return "—"
    if v >= green:
        return "Green"
    if v >= yellow:
        return "Yellow"
    return "Red"


def farol_low(v, green, yellow):
    if v is None:
        return "—"
    if v <= green:
        return "Green"
    if v <= yellow:
        return "Yellow"
    return "Red"


def paint_farol(cell, label):
    if label == "Green":
        cell.fill = C_GREEN
    elif label == "Yellow":
        cell.fill = C_YELLOW
    elif label == "Red":
        cell.fill = C_RED
    else:
        cell.fill = C_GRAY


def write_matrix(wb, sheet_name, title, buckets, periods, value_fn, sample_fn, fmt,
                 green, yellow, lower_is_better=False):
    ws = wb.create_sheet(sheet_name)
    # layout: col A = TSA, then pairs (period / sample)
    n_periods = len(periods)
    total_cols = 1 + n_periods * 2 + 3  # + Média + Amostra Total + Farol
    write_title(ws, title, total_cols)

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 24

    # period mega-header (row 3)
    ws.cell(row=3, column=1, value="TSA").fill = C_SUB
    ws.cell(row=3, column=1).font = F_H
    ws.cell(row=3, column=1).alignment = CENTER
    ws.cell(row=3, column=1).border = BORDER
    ws.cell(row=4, column=1, value="").border = BORDER

    col = 2
    for p in periods:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        c = ws.cell(row=3, column=col, value=p)
        c.fill = C_SUB
        c.font = F_H
        c.alignment = CENTER
        c.border = BORDER

        sub_val = ws.cell(row=4, column=col, value="Value")
        sub_n = ws.cell(row=4, column=col + 1, value="n")
        for s in (sub_val, sub_n):
            s.fill = C_GRAY
            s.font = F_BODY_B
            s.alignment = CENTER
            s.border = BORDER
        col += 2

    # Totals headers
    for idx, label in enumerate(["Overall avg", "Total sample", "Status"]):
        ws.merge_cells(start_row=3, start_column=col + idx, end_row=4, end_column=col + idx)
        c = ws.cell(row=3, column=col + idx, value=label)
        c.fill = C_HEADER
        c.font = F_H
        c.alignment = CENTER
        c.border = BORDER

    # body
    for i, tsa in enumerate(KPI_MEMBERS, start=5):
        c0 = ws.cell(row=i, column=1, value=tsa)
        c0.font = F_BODY_B
        c0.fill = C_CARD
        c0.alignment = CENTER
        c0.border = BORDER

        col = 2
        global_vals = []
        global_n = 0
        for p in periods:
            entry = buckets.get((tsa, p))
            if entry is None:
                val = None
                n = 0
            else:
                val = value_fn(entry)
                n = sample_fn(entry)

            cv = ws.cell(row=i, column=col, value=val if val is not None else "")
            cn = ws.cell(row=i, column=col + 1, value=n)
            if val is not None:
                cv.number_format = fmt
                global_vals.append((val, n))
                global_n += n
                # paint cell by farol
                if lower_is_better:
                    label = farol_low(val, green, yellow)
                else:
                    label = farol_pct(val, green, yellow)
                paint_farol(cv, label)
            else:
                cv.fill = C_GRAY
            cv.alignment = CENTER
            cn.alignment = CENTER
            cv.border = BORDER
            cn.border = BORDER
            cv.font = F_BODY
            cn.font = F_BODY
            col += 2

        # Totals per TSA
        if global_vals:
            if lower_is_better:
                # weighted by n
                total_days = sum(v * n for v, n in global_vals)
                mean = total_days / global_n if global_n else None
            else:
                # for percentages, re-derive from underlying n — but we only have (val, n),
                # so recompute as weighted mean
                total_num = sum(v * n for v, n in global_vals)
                mean = total_num / global_n if global_n else None
        else:
            mean = None

        ws.cell(row=i, column=col, value=mean if mean is not None else "").number_format = fmt
        ws.cell(row=i, column=col).alignment = CENTER
        ws.cell(row=i, column=col).font = F_BODY_B
        ws.cell(row=i, column=col).border = BORDER

        ws.cell(row=i, column=col + 1, value=global_n).alignment = CENTER
        ws.cell(row=i, column=col + 1).font = F_BODY_B
        ws.cell(row=i, column=col + 1).border = BORDER

        label = (farol_low(mean, green, yellow) if lower_is_better else farol_pct(mean, green, yellow))
        cf = ws.cell(row=i, column=col + 2, value=label)
        cf.alignment = CENTER
        cf.font = F_BODY_B
        cf.border = BORDER
        paint_farol(cf, label)

    # column widths
    ws.column_dimensions["A"].width = 14
    for c in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # freeze
    ws.freeze_panes = "B5"


def write_dados(wb, records):
    ws = wb.create_sheet("Raw Data")
    cols = ["tsa", "period", "week", "weekRange", "customer", "category",
            "status", "perf", "rework", "dateAdd", "startedAt", "eta",
            "deliveryDate", "delivery", "velocity_days", "source", "ticketId", "ticketUrl", "focus"]
    write_title(ws, "Raw data (same filter as live dashboard: Linear + External + core weeks)",
                len(cols))

    for j, h in enumerate(cols, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.fill = C_SUB
        c.font = F_H
        c.alignment = CENTER
        c.border = BORDER

    row = 4
    for r in records:
        sd = parse_date(r.get("startedAt", ""))
        dd = parse_date(r.get("deliveryDate", ""))
        vel = (dd - sd).days if sd and dd and (dd - sd).days >= 0 else ""
        values = [
            r.get("tsa", ""),
            period_from_week(r.get("week", "")),
            r.get("week", ""),
            r.get("weekRange", ""),
            r.get("customer", ""),
            r.get("category", ""),
            r.get("status", ""),
            r.get("perf", ""),
            "Yes" if r.get("rework") else "",
            r.get("dateAdd", ""),
            r.get("startedAt", "")[:10],
            r.get("eta", ""),
            r.get("deliveryDate", "")[:10],
            r.get("delivery", ""),
            vel,
            r.get("source", ""),
            r.get("ticketId", ""),
            r.get("ticketUrl", ""),
            (r.get("focus", "") or "")[:120],
        ]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=j, value=v)
            c.font = F_BODY
            c.alignment = LEFT if j in (5, 19) else CENTER
            c.border = BORDER
        row += 1

    widths = [11, 10, 12, 20, 16, 10, 14, 12, 9, 12, 12, 12, 12, 12, 10, 11, 14, 40, 40]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}{row - 1}"


# ---------------- Main ----------------

def write_onboarding(wb, timeline):
    """Indicator 1 — Customer Onboarding. Table + portfolio summary."""
    ws = wb.create_sheet("KPI1 Customer Onboarding")
    cols = ["Customer", "Deal Signed", "Kickoff", "Go-Live", "Duration (mo)", "Status", "Status Class"]
    write_title(ws, "KPI 1 — Customer Onboarding (kickoff → go-live) — Target ≤ 3 months",
                len(cols))

    for j, h in enumerate(cols, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.fill = C_SUB
        c.font = F_H
        c.alignment = CENTER
        c.border = BORDER

    status_class = {
        "live": "Live", "maintenance": "Maintenance",
        "in_progress": "In Progress", "starting": "Starting",
        "interrupted": "Interrupted", "stalled": "Stalled",
        "churned": "Churned", "pre_sales": "Pre-sales",
    }
    status_order = {"live": 0, "maintenance": 1, "in_progress": 2, "starting": 3,
                    "interrupted": 4, "stalled": 5, "pre_sales": 6, "churned": 7}

    # Compute months per project + sort
    rows_data = []
    for r in timeline:
        kd = parse_mon_yy(r.get("kickoff"))
        gd = parse_mon_yy(r.get("goLive"))
        months = months_diff(kd, gd) if kd and gd else None
        if months is None and kd and r.get("status") in ("in_progress", "starting"):
            today = date.today()
            months_ongoing = (today.year - kd.year) * 12 + (today.month - kd.month)
            dur_str = f"{months_ongoing}+"
        elif months is not None:
            dur_str = months
        else:
            dur_str = "—"
        rows_data.append({
            "customer": r.get("customer", ""),
            "dealSigned": r.get("dealSigned", "") or "—",
            "kickoff": r.get("kickoff", "") or "—",
            "goLive": r.get("goLive", "") or "—",
            "months": months,
            "dur_str": dur_str,
            "status": r.get("status", ""),
            "kn": yearmonth_num(kd) if kd else 999999,
        })

    rows_data.sort(key=lambda x: (status_order.get(x["status"], 9), x["kn"]))

    row = 4
    for r in rows_data:
        vals = [r["customer"], r["dealSigned"], r["kickoff"], r["goLive"],
                r["dur_str"], r["status"], status_class.get(r["status"], r["status"])]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=j, value=v)
            c.font = F_BODY_B if j == 1 else F_BODY
            c.alignment = LEFT if j == 1 else CENTER
            c.border = BORDER
            if j == 5 and isinstance(r["months"], int):
                if r["months"] <= ONB_GREEN:
                    c.fill = C_GREEN
                elif r["months"] <= ONB_YELLOW:
                    c.fill = C_YELLOW
                else:
                    c.fill = C_RED
            if j == 7:
                if r["status"] in ("live", "maintenance"):
                    c.fill = C_GREEN
                elif r["status"] in ("in_progress", "starting"):
                    c.fill = C_YELLOW
                elif r["status"] in ("stalled", "interrupted", "churned"):
                    c.fill = C_RED
                else:
                    c.fill = C_GRAY
        row += 1

    # Portfolio summary block
    ws.cell(row=row + 1, column=1, value="Portfolio Summary").font = F_BODY_B
    row += 2

    completed = [r for r in rows_data if r["months"] is not None]
    avg_months = sum(r["months"] for r in completed) / len(completed) if completed else None
    # TSA Team era: kickoff ≥ Sep/25
    tsa_completed = [r for r in completed
                     if (r["kn"] != 999999 and r["kn"] >= TSA_ERA_YEARMONTH)
                     or r["customer"] == "WFS"]
    tsa_avg = sum(r["months"] for r in tsa_completed) / len(tsa_completed) if tsa_completed else None
    fastest = min(completed, key=lambda r: r["months"]) if completed else None

    def status_count(s):
        if isinstance(s, list):
            return sum(1 for r in rows_data if r["status"] in s)
        return sum(1 for r in rows_data if r["status"] == s)

    summary = [
        ("Avg onboarding time", f"{avg_months:.1f} mo" if avg_months else "—"),
        ("TSA Team Avg (since Sep/25)", f"{tsa_avg:.1f} mo" if tsa_avg else "—"),
        ("Fastest go-live", f"{fastest['customer']} — {fastest['months']} mo" if fastest else "—"),
        ("Live / Maintenance", status_count(["live", "maintenance"])),
        ("In Progress / Starting", status_count(["in_progress", "starting"])),
        ("Interrupted / Stalled", status_count(["interrupted", "stalled"])),
        ("Churned", status_count("churned")),
    ]
    for k, v in summary:
        ws.cell(row=row, column=1, value=k).font = F_BODY_B
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=2, value=v).alignment = LEFT
        ws.cell(row=row, column=2).font = F_BODY
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A4"


def main():
    all_records = load()
    records = filter_kpi(all_records)  # Dashboard parity: Linear + External + core weeks
    timeline = load_timeline()

    k1 = compute_kpi1(records)   # ETA Accuracy (TSA KPI)
    k3 = compute_kpi3(records)   # Reliability

    periods = periods_sorted(k1, k3, n_last=12)

    print(f"All records: {len(all_records)} | Dashboard-filtered: {len(records)}")
    print(f"Filters applied: source=linear, category=External, core weeks (rolling {CORE_WEEK_ROLLING_MONTHS} months)")
    print(f"Timeline projects: {len(timeline)}")

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    write_cover(wb)
    write_resumo(wb, k1, k3, periods, timeline)

    # Indicator 1 — Customer Onboarding (per-customer project view)
    write_onboarding(wb, timeline)

    # Indicator 2 — ETA Accuracy (per-TSA per-period)
    write_matrix(
        wb, "KPI2 ETA Accuracy",
        "KPI 2 — ETA Accuracy (On Time / On Time + Late) — Target ≥ 90%",
        k1, periods,
        value_fn=lambda e: (e["on_time"] / (e["on_time"] + e["late"])) if (e["on_time"] + e["late"]) else None,
        sample_fn=lambda e: e["on_time"] + e["late"],
        fmt="0.0%",
        green=KPI1_GREEN, yellow=KPI1_YELLOW, lower_is_better=False,
    )

    # Indicator 3 — Implementation Reliability (per-TSA per-period)
    write_matrix(
        wb, "KPI3 Reliability",
        "KPI 3 — Implementation Reliability (Done without Rework / Done) — Target ≥ 90%",
        k3, periods,
        value_fn=lambda e: ((e["done"] - e["rework"]) / e["done"]) if e["done"] else None,
        sample_fn=lambda e: e["done"],
        fmt="0.0%",
        green=KPI3_GREEN, yellow=KPI3_YELLOW, lower_is_better=False,
    )

    write_dados(wb, records)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    target = OUT
    try:
        wb.save(target)
    except PermissionError:
        target = OUT.with_name(OUT.stem + "_v2.xlsx")
        wb.save(target)
        print(f"WARN: {OUT.name} was locked (likely open in Excel). Saved as {target.name} instead.")
    print(f"OK: {target}  ({target.stat().st_size // 1024} KB)")
    print(f"Periods used: {periods}")
    print(f"KPI1 Onboarding projects: {len(timeline)} | KPI2 ETA buckets: {len(k1)} | KPI3 Rel buckets: {len(k3)}")


if __name__ == "__main__":
    main()
