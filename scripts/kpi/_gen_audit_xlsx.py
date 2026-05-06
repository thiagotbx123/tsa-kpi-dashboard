"""Generate KPI Pipeline Re-Audit XLSX — auto-dated output to Downloads."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

_DATE = datetime.now().strftime('%Y-%m-%d')
OUTPUT = rf'C:\Users\adm_r\Downloads\AUDIT_KPI_PIPELINE_{_DATE}.xlsx'

CRIT_FILL  = PatternFill('solid', fgColor='DC2626')
HIGH_FILL  = PatternFill('solid', fgColor='F97316')
MED_FILL   = PatternFill('solid', fgColor='FDE047')
LOW_FILL   = PatternFill('solid', fgColor='BAE6FD')
HDR_FILL   = PatternFill('solid', fgColor='1E293B')
GREEN_FILL = PatternFill('solid', fgColor='DCFCE7')
FIXED_FILL = PatternFill('solid', fgColor='DCFCE7')
NEW_FILL   = PatternFill('solid', fgColor='FEF3C7')

HDR_FONT   = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
WHITE_FONT = Font(bold=True, color='FFFFFF', name='Calibri')
BOLD_FONT  = Font(bold=True, name='Calibri')
BASE_FONT  = Font(name='Calibri')
TITLE_FONT = Font(bold=True, size=16, name='Calibri', color='1E293B')
H2_FONT    = Font(bold=True, size=12, name='Calibri', color='1E293B')
SCORE_FONT = Font(bold=True, size=22, name='Calibri', color='16A34A')

thin = Side(style='thin', color='CBD5E1')
BORDER = Border(top=thin, left=thin, bottom=thin, right=thin)

SEV_FILL = {'CRITICAL': CRIT_FILL, 'HIGH': HIGH_FILL, 'MEDIUM': MED_FILL, 'LOW': LOW_FILL}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# RE-AUDIT FINDINGS — 2026-04-02 post-fixes
# Original audit: 68/100 YELLOW (1C, 8H, 15M, 16L = 40 total)
# Re-audit:       82/100 GREEN  (0C, 2H, 12M, 8L  = 22 total)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

FINDINGS = [
    # --- 2 HIGH remain ---
    ('A03-001', 'HIGH', 'CODE-QUALITY',
     'build_html_dashboard.py is 2838-line monolith with embedded HTML/CSS/JS as raw Python string.',
     'build_html_dashboard.py',
     'Extract HTML template to separate .html file with Jinja2 or mako.',
     'HIGH'),
    ('A08-001', 'HIGH', 'TESTING',
     '6 of 8 core modules have zero test coverage. Only merge + normalize pure functions tested.',
     'tests/test_kpi_calculations.py',
     'Add integration tests for orchestrate, build, refresh, upload, tray, serve.',
     'HIGH'),

    # --- 12 MEDIUM remain ---
    ('A03-002', 'MEDIUM', 'CODE-QUALITY',
     'Module-level execution in merge/normalize prevents clean imports. No if __name__=="__main__" guard.',
     'merge_opossum_data.py / normalize_data.py',
     'Wrap main logic in def main(); export pure functions.',
     'HIGH'),
    ('A01-003', 'MEDIUM', 'DATA-INTEGRITY',
     'State IDs (15 UUIDs) hardcoded without runtime validation against Linear API. Mitigated: unknown IDs now logged.',
     'merge_opossum_data.py:24-44',
     'Add startup validation or fallback for unknown states.',
     'MEDIUM'),
    ('A09-002', 'MEDIUM', 'BUSINESS-LOGIC',
     'Admin-close detection may misclassify real tickets. Done + empty startedAt forced to On Time.',
     'normalize_data.py:339-353',
     'Add guard: check history event count >= 2 before treating as admin-close.',
     'MEDIUM'),
    ('A11-001', 'MEDIUM', 'DATA-INTEGRITY',
     'Partial pipeline failure leaves _dashboard_data.json with unrecalculated perf labels.',
     'orchestrate.py',
     'Write merge output to staging file; promote after normalize succeeds.',
     'MEDIUM'),
    ('A11-002', 'MEDIUM', 'DATA-INTEGRITY',
     'Cache freshness 24h warning is print-only. Dashboard can be built from week-old data.',
     'merge_opossum_data.py',
     'Pass cache age to dashboard; show red banner if >48h.',
     'MEDIUM'),
    ('A14-001', 'MEDIUM', 'INFRASTRUCTURE',
     'No CI/CD pipeline. No automated testing before deployment.',
     'Project root',
     'Add GitHub Actions: pytest on push to master.',
     'MEDIUM'),
    ('A17-001', 'MEDIUM', 'ACCESSIBILITY',
     'Dashboard tabs are clickable divs, not button/tab elements. Chart interactions mouse-only.',
     'build_html_dashboard.py JS',
     'Add role=tab, aria-selected; add data table alternative.',
     'MEDIUM'),
    ('A20-001', 'MEDIUM', 'MAINTAINABILITY',
     'Bus factor = 1. Single contributor. 59+ data integrity rules with subtle interactions.',
     'Git history',
     'Document top 10 maintenance tasks. Add troubleshooting guide.',
     'HIGH'),
    ('A27-001', 'MEDIUM', 'BUSINESS-LOGIC',
     'ETA gaming: set ETA = delivery date guarantees On Time. retroactiveEta flag informational only.',
     'merge_opossum_data.py',
     'Show "Organic ETA Accuracy" excluding retroactive ETAs alongside total.',
     'HIGH'),
    ('A32-001', 'MEDIUM', 'BUSINESS-LOGIC',
     'No validation of state transition legality. Unusual paths (Triage->Done) not flagged.',
     'merge_opossum_data.py',
     'Flag tickets with skipped In Progress step.',
     'MEDIUM'),
    ('A35-002', 'MEDIUM', 'BUSINESS-LOGIC',
     'Velocity counts tickets equally regardless of estimate/complexity.',
     'Dashboard JS',
     'Show weighted velocity (sum of estimates) alongside count.',
     'MEDIUM'),
    ('A29-001', 'MEDIUM', 'INFRASTRUCTURE',
     'Python 3.14 is bleeding-edge pre-release. Library compat issues possible.',
     'kpi_publish.bat',
     'Test with stable 3.12/3.13. Pin in docs.',
     'MEDIUM'),

    # --- 8 LOW remain ---
    ('A03-003', 'LOW', 'CODE-QUALITY',
     '10 variant files in variants/ (5534 lines) appear unused. Dead code.',
     'variants/',
     'Move to archive/ or delete.',
     'HIGH'),
    ('A19-002', 'LOW', 'ARCHITECTURE',
     'Two HTTP servers for same purpose: serve_kpi.py (8787) and kpi_tray.py (8080).',
     'serve_kpi.py / kpi_tray.py',
     'Deprecate serve_kpi.py -- tray is canonical.',
     'LOW'),
    ('N01', 'LOW', 'ERROR-HANDLING',
     'Upload retry doesn\'t catch JSONDecodeError on non-JSON error bodies (429/5xx with HTML).',
     'upload_dashboard_drive.py',
     'Wrap resp.json() in try/except JSONDecodeError.',
     'MEDIUM'),
    ('N02', 'LOW', 'SECURITY',
     'ngrok basic-auth credentials hardcoded in source. Anyone with repo access sees them.',
     'kpi_tray.py:219',
     'Move credentials to .env file or OS keyring.',
     'MEDIUM'),
    ('N03', 'LOW', 'CODE-QUALITY',
     'normalize calc_perf and calc_perf_with_history duplicate status gate logic. Drift risk.',
     'normalize_data.py:71-150',
     'Extract shared status gates into helper function.',
     'LOW'),
    ('N04', 'LOW', 'CODE-QUALITY',
     'kpi_tray _LOG_DIR hardcoded to ~/Downloads/kpi-serve instead of using OUTPUT_DIR from team_config.',
     'kpi_tray.py',
     'Import OUTPUT_DIR; derive _LOG_DIR from it.',
     'LOW'),
    ('N05', 'LOW', 'DATA-INTEGRITY',
     'D.LIE22 dedup: bracket strip only on Linear side; spreadsheet keys don\'t strip. Mismatch possible.',
     'normalize_data.py:377-387',
     'Normalize both sides consistently.',
     'LOW'),
    ('N06', 'LOW', 'CODE-QUALITY',
     'Hardcoded 2019->2025 year repair in normalize. May break for other years or ongoing bad data.',
     'normalize_data.py:190-206',
     'Make year repair configurable or remove if no longer needed.',
     'LOW'),
]

WORKING = [
    ('I01', 'Linear API integration with pagination works reliably for standard issue fetch'),
    ('I02', 'History extraction correctly captures ETA changes, state transitions, assignee changes'),
    ('I03', 'D.LIE scoring engine covers 6 distinct lie dimensions with clear threshold logic'),
    ('I04', 'Gantt chart rendering handles multi-week spans and color-codes by performance tier'),
    ('I05', 'Filter system (person / week / state) correctly narrows all computed tables'),
    ('I06', 'Heatmap generation correctly aggregates weekly on-time rates per person'),
    ('I07', 'orchestrate.py provides single-command full-pipeline execution'),
    ('I08', 'kpi_tray.py system tray with one-click refresh + serve + ngrok (now with auth)'),
    ('I09', 'Cache-first architecture reduces API calls and allows offline dashboard viewing'),
    ('I10', 'Collapse sections provide clean progressive disclosure of detail rows'),
    ('I11', 'Comprehensive 59-rule D.LIE integrity framework well-documented'),
    ('I12', 'Dual-mode tray (full_refresh vs skip_refresh) enables fast dev iteration'),
    ('I13', 'team_config.py serves as SSOT for person maps, customer maps, perf constants, OUTPUT_DIR'),
    ('I14', 'normalize_data.py is single authority for perf calc (A30-003 resolved)'),
    ('I15', 'Upload to Drive has retry with exponential backoff (3 attempts)'),
    ('I16', 'Dashboard shows API cache mtime alongside build date for accurate freshness'),
    ('I17', 'HTTP server bound to localhost only; ngrok has basic auth for external access'),
    ('I18', 'Log rotation via RotatingFileHandler (5MB, 3 backups) in tray server'),
]

ACTIONS = [
    ('1', 'A03-001',          'CODE-QUALITY',    'Extract HTML template from build_html_dashboard.py monolith.'),
    ('2', 'A08-001',          'TESTING',          'Add integration tests for orchestrate, build, refresh, upload modules.'),
    ('3', 'A03-002',          'CODE-QUALITY',     'Wrap merge/normalize main logic in def main(); add __name__ guard.'),
    ('4', 'A09-002+A27-001',  'BUSINESS-LOGIC',   'Add admin-close guards and "Organic ETA Accuracy" metric.'),
    ('5', 'A14-001',          'INFRASTRUCTURE',   'Add GitHub Actions CI: pytest on push to master.'),
    ('6', 'A20-001',          'MAINTAINABILITY',  'Write troubleshooting guide and maintenance runbook.'),
    ('7', 'A17-001',          'ACCESSIBILITY',    'Add ARIA roles to dashboard tabs; data table alternative.'),
    ('8', 'N01+N02',          'ERROR-HANDLING',   'Fix upload JSONDecodeError gap; move ngrok creds to .env.'),
    ('9', 'A03-003',          'CODE-QUALITY',     'Clean up dead code in variants/ (5534 lines unused).'),
    ('10', 'A32-001+A35-002', 'BUSINESS-LOGIC',   'State transition validation + weighted velocity metric.'),
]

CAT_SCORES = [
    ('Security',        75, 100, '16A34A'),
    ('Data Integrity',  78, 100, '16A34A'),
    ('Error Handling',  72, 100, '16A34A'),
    ('Testing',         40, 100, 'F97316'),
    ('Dependencies',    85, 100, '16A34A'),
    ('Business Logic',  65, 100, 'FDE047'),
    ('Code Quality',    62, 100, 'FDE047'),
    ('Infrastructure',  70, 100, '16A34A'),
    ('Configuration',   88, 100, '16A34A'),
    ('Accessibility',   45, 100, 'F97316'),
    ('Documentation',   75, 100, '16A34A'),
    ('Maintainability', 58, 100, 'F97316'),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# REGRESSION DELTA — Original audit → Re-audit
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

REGRESSION = [
    ('FIXED', 'A30-003', 'CRITICAL: Dual calc_perf ambiguity. merge now uses _placeholder_perf; normalize is sole authority.'),
    ('FIXED', 'A06-001', 'HIGH: Phantom deps (pystray, Pillow, google-auth, google-auth-oauthlib) added to requirements.txt.'),
    ('FIXED', 'A04-003', 'HIGH: HTTP server now binds to 127.0.0.1 (localhost only).'),
    ('FIXED', 'A04-004', 'HIGH: ngrok now has --basic-auth for tunnel access.'),
    ('FIXED', 'A01-002', 'HIGH: Customer/person maps consolidated into team_config.py single source of truth.'),
    ('FIXED', 'A37-002', 'HIGH: Dashboard now shows API cache mtime + build date (was build date only).'),
    ('FIXED', 'A21-002', 'HIGH: Filename unified to KPI_DASHBOARD.html everywhere (build, upload, serve, docstrings).'),
    ('FIXED', 'A07-001', 'MEDIUM: .env.example created with LINEAR_API_KEY and Google OAuth setup.'),
    ('FIXED', 'A07-002', 'MEDIUM: OUTPUT_DIR centralized in team_config.py; all modules import it.'),
    ('FIXED', 'A02-003', 'MEDIUM: Upload to Drive now has retry with exponential backoff (3 attempts).'),
    ('FIXED', 'A13-002', 'MEDIUM: Log rotation via RotatingFileHandler (5MB, 3 backups) properly wired in kpi_tray.'),
    ('FIXED', 'A01-004', 'LOW: Bare except replaced with specific (json.JSONDecodeError, IOError, ValueError).'),
    ('FIXED', 'A03-004', 'LOW: Perf label constants (PERF_ON_TIME, PERF_LATE, etc.) in team_config; used everywhere.'),
    ('FIXED', '—',       'spike:None → spike:Internal in CUSTOMER_MAP (prevented None propagation).'),
    ('FIXED', '—',       'REAL_CUSTOMERS de-duped: removed "BILL", kept "Bill".'),
    ('FIXED', '—',       'Unused REAL_CUSTOMERS import removed from normalize_data.py.'),
    ('FIXED', '—',       'serve_kpi.py docstring corrected (RACCOONS_ → KPI_DASHBOARD).'),
    ('FIXED', '—',       '_placeholder_perf now uses PERF_* constants instead of string literals.'),
    ('OPEN',  'A03-001', 'HIGH: build_html_dashboard.py still 2838-line monolith (architectural debt).'),
    ('OPEN',  'A08-001', 'HIGH: 6/8 modules still lack test coverage.'),
    ('OPEN',  'A03-002', 'MEDIUM: merge/normalize still run at module level (no __main__ guard).'),
    ('NEW',   'N01',     'Upload retry gap: resp.json() on non-JSON bodies raises JSONDecodeError.'),
    ('NEW',   'N02',     'ngrok creds hardcoded in source (kpi_tray.py:219).'),
    ('NEW',   'N03',     'calc_perf / calc_perf_with_history duplicated status gates inside normalize.'),
    ('NEW',   'N04',     'kpi_tray _LOG_DIR not derived from OUTPUT_DIR.'),
    ('NEW',   'N05',     'D.LIE22 bracket strip asymmetry between Linear and spreadsheet keys.'),
    ('NEW',   'N06',     'Hardcoded 2019->2025 year repair may need update.'),
]


def build_findings_sheet(wb):
    ws = wb.active
    ws.title = 'Findings'
    ws.freeze_panes = 'A2'

    headers = ['ID', 'Severity', 'Category', 'Summary', 'File:Line', 'Recommendation', 'Confidence']
    col_widths = [10, 12, 18, 60, 34, 55, 12]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws.row_dimensions[1].height = 20

    for r, f in enumerate(FINDINGS, 2):
        sev = f[1]
        fill = SEV_FILL.get(sev)
        for col, val in enumerate(f, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.font = BASE_FONT
        sev_cell = ws.cell(row=r, column=2)
        if fill:
            sev_cell.fill = fill
        if sev == 'CRITICAL':
            sev_cell.font = WHITE_FONT
        elif sev in ('HIGH', 'MEDIUM', 'LOW'):
            sev_cell.font = BOLD_FONT
        for col in (1, 3, 4, 5, 6, 7):
            cell = ws.cell(row=r, column=col)
            if r % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F8FAFC')

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for r in range(2, len(FINDINGS) + 2):
        ws.row_dimensions[r].height = 44


def build_summary_sheet(wb):
    ws = wb.create_sheet('Summary')

    def cell(row, col, value, font=None, fill=None, align=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:  c.font = font
        if fill:  c.fill = fill
        if align: c.alignment = align
        return c

    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1, value='KPI Pipeline RE-AUDIT (AUDIT_ENGINE v3.2)')
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor='F1F5F9')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:G2')
    c2 = ws.cell(row=2, column=1,
                 value=f'Date: {_DATE}  |  Target: Tools/TSA_CORTEX/scripts/kpi/  |  Post-fix re-assessment  |  Tests: 88/88 PASS')
    c2.font = Font(name='Calibri', size=10, color='64748B')
    c2.fill = PatternFill('solid', fgColor='F1F5F9')

    ws.row_dimensions[4].height = 28
    cell(4, 1, 'OVERALL HEALTH SCORE', font=H2_FONT)
    cell(4, 2, '82 / 100', font=SCORE_FONT,
         align=Alignment(horizontal='center', vertical='center'))
    cell(4, 3, 'GREEN', font=Font(bold=True, color='16A34A', name='Calibri', size=13),
         align=Alignment(horizontal='center', vertical='center'))
    ws.merge_cells('D4:G4')
    cell(4, 4, 'Up from 68/100 YELLOW. CRITICAL eliminated, 6/8 HIGHs fixed. Remaining: monolith + test coverage.',
         font=Font(italic=True, name='Calibri', color='475569'))

    cell(6, 1, 'SEVERITY BREAKDOWN', font=H2_FONT)
    hdr_row = 7
    for col, h in enumerate(['Severity', 'Before', 'After', 'Delta'], 1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center'); c.border = BORDER

    before = [('CRITICAL', 1), ('HIGH', 8), ('MEDIUM', 15), ('LOW', 16)]
    after_counts = {
        'CRITICAL': sum(1 for f in FINDINGS if f[1] == 'CRITICAL'),
        'HIGH':     sum(1 for f in FINDINGS if f[1] == 'HIGH'),
        'MEDIUM':   sum(1 for f in FINDINGS if f[1] == 'MEDIUM'),
        'LOW':      sum(1 for f in FINDINGS if f[1] == 'LOW'),
    }
    sev_fills = {'CRITICAL': (CRIT_FILL, WHITE_FONT), 'HIGH': (HIGH_FILL, BOLD_FONT),
                 'MEDIUM': (MED_FILL, BOLD_FONT), 'LOW': (LOW_FILL, BOLD_FONT)}

    for i, (sev, bcount) in enumerate(before, hdr_row + 1):
        acount = after_counts[sev]
        delta = acount - bcount
        sfill, sfont = sev_fills[sev]
        c1 = ws.cell(row=i, column=1, value=sev)
        c1.fill = sfill; c1.font = sfont; c1.border = BORDER; c1.alignment = Alignment(horizontal='center')
        c2 = ws.cell(row=i, column=2, value=bcount)
        c2.font = BOLD_FONT; c2.border = BORDER; c2.alignment = Alignment(horizontal='center')
        c3 = ws.cell(row=i, column=3, value=acount)
        c3.font = BOLD_FONT; c3.border = BORDER; c3.alignment = Alignment(horizontal='center')
        c4 = ws.cell(row=i, column=4, value=f'{delta:+d}')
        c4.border = BORDER; c4.alignment = Alignment(horizontal='center')
        c4.font = Font(bold=True, color='16A34A' if delta < 0 else 'DC2626' if delta > 0 else '475569', name='Calibri')

    total_row = hdr_row + len(before) + 1
    ws.cell(row=total_row, column=1, value='TOTAL').font = BOLD_FONT
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=total_row, column=2, value=40).font = BOLD_FONT
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=total_row, column=3, value=len(FINDINGS)).font = BOLD_FONT
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='center')
    delta_total = len(FINDINGS) - 40
    c_dt = ws.cell(row=total_row, column=4, value=f'{delta_total:+d}')
    c_dt.font = Font(bold=True, color='16A34A', name='Calibri')
    c_dt.alignment = Alignment(horizontal='center')
    for col in range(1, 5):
        ws.cell(row=total_row, column=col).border = BORDER

    cat_start = total_row + 2
    cell(cat_start, 1, 'CATEGORY SCORES', font=H2_FONT)
    hcat = cat_start + 1
    for col, h in enumerate(['Category', 'Score', 'Out of', 'Grade'], 1):
        c = ws.cell(row=hcat, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center'); c.border = BORDER

    for i, (cat, score, out_of, color) in enumerate(CAT_SCORES, hcat + 1):
        pct = score / out_of
        grade = 'A' if pct >= .85 else 'B' if pct >= .7 else 'C' if pct >= .6 else 'D' if pct >= .5 else 'F'
        cfill = PatternFill('solid', fgColor=color)
        ws.cell(row=i, column=1, value=cat).border = BORDER
        ws.cell(row=i, column=2, value=score).border = BORDER
        ws.cell(row=i, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=3, value=out_of).border = BORDER
        ws.cell(row=i, column=3).alignment = Alignment(horizontal='center')
        gc = ws.cell(row=i, column=4, value=grade)
        gc.border = BORDER; gc.fill = cfill
        gc.alignment = Alignment(horizontal='center')
        if color in ('DC2626', 'F97316'):
            gc.font = Font(bold=True, color='FFFFFF', name='Calibri')
        else:
            gc.font = BOLD_FONT

    act_start = hcat + len(CAT_SCORES) + 2
    cell(act_start, 1, 'TOP 10 ACTION ITEMS (remaining)', font=H2_FONT)
    hact = act_start + 1
    for col, h in enumerate(['#', 'Finding(s)', 'Category', 'Action'], 1):
        c = ws.cell(row=hact, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center'); c.border = BORDER

    for i, (num, refs, cat, action) in enumerate(ACTIONS, hact + 1):
        ws.cell(row=i, column=1, value=num).border = BORDER
        ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')
        rc = ws.cell(row=i, column=2, value=refs)
        rc.border = BORDER
        rc.font = Font(color='4F46E5', bold=True, name='Calibri')
        rc.alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=3, value=cat).border = BORDER
        ac = ws.cell(row=i, column=4, value=action)
        ac.border = BORDER; ac.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 30

    good_start = hact + len(ACTIONS) + 2
    cell(good_start, 1, "WHAT'S WORKING WELL (expanded after fixes)", font=H2_FONT)
    hgood = good_start + 1
    for col, h in enumerate(['ID', 'Observation'], 1):
        c = ws.cell(row=hgood, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.border = BORDER; c.alignment = Alignment(horizontal='center')

    for i, (iid, obs) in enumerate(WORKING, hgood + 1):
        ic = ws.cell(row=i, column=1, value=iid)
        ic.fill = GREEN_FILL; ic.border = BORDER
        ic.font = Font(bold=True, color='166534', name='Calibri')
        ic.alignment = Alignment(horizontal='center')
        oc = ws.cell(row=i, column=2, value=obs)
        oc.fill = GREEN_FILL; oc.border = BORDER
        oc.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 22

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 65
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12


def build_regression_sheet(wb):
    ws = wb.create_sheet('Regression Delta')
    ws.freeze_panes = 'A2'

    OPEN_FILL = PatternFill('solid', fgColor='FEE2E2')

    headers = ['Status', 'Finding ID', 'Description']
    col_widths = [10, 14, 80]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

    for r, (status, fid, desc) in enumerate(REGRESSION, 2):
        if status == 'FIXED':
            fill = FIXED_FILL
            sfont = Font(bold=True, color='166534', name='Calibri')
        elif status == 'OPEN':
            fill = OPEN_FILL
            sfont = Font(bold=True, color='991B1B', name='Calibri')
        else:
            fill = NEW_FILL
            sfont = Font(bold=True, color='92400E', name='Calibri')
        c1 = ws.cell(row=r, column=1, value=status)
        c1.fill = fill; c1.font = sfont; c1.border = BORDER
        c1.alignment = Alignment(horizontal='center')
        c2 = ws.cell(row=r, column=2, value=fid)
        c2.fill = fill; c2.border = BORDER; c2.font = BOLD_FONT
        c2.alignment = Alignment(horizontal='center')
        c3 = ws.cell(row=r, column=3, value=desc)
        c3.fill = fill; c3.border = BORDER
        c3.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 22

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.row_dimensions[1].height = 20


wb = openpyxl.Workbook()
build_findings_sheet(wb)
build_summary_sheet(wb)
build_regression_sheet(wb)

wb.save(OUTPUT)
print(f'Saved: {OUTPUT}')
print(f'Findings: {len(FINDINGS)} | Strengths: {len(WORKING)} | Actions: {len(ACTIONS)} | Regression: {len(REGRESSION)}')
fixed_count = sum(1 for r in REGRESSION if r[0] == 'FIXED')
print(f'Regression: {fixed_count} FIXED, {len(REGRESSION) - fixed_count} OPEN/NEW')
